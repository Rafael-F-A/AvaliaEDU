import csv
import io
import secrets
import string
from typing import Any

import openpyxl
from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session

from app import models
from app.security import hash_senha

# Constantes

NIVEIS_VALIDOS = {"FUNDAMENTAL_I", "FUNDAMENTAL_II", "MEDIO", "ENEM", "EJA"}
TAMANHO_MAXIMO = 5 * 1024 * 1024  # 5 MB
CONTENT_TYPES_PERMITIDOS = {
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    # alguns browsers enviam text/plain para CSV
    "text/plain",
}

# Mapeamento flexível de cabeçalhos (aceita variações em PT e EN)
MAPA_COLUNAS = {
    "nome":   ["nome", "name", "aluno", "student"],
    "email":  ["email", "e-mail", "e_mail", "mail"],
    "nivel":  ["nivel", "nível", "level", "nivel_ensino"],
    "serie":  ["serie", "série", "series", "turma", "class"],
    "senha":  ["senha", "password"],  # opcional — ignorado se não vier
}

# Helpers

def _gerar_senha_provisoria() -> str:
    """Gera senha aleatória de 12 chars para o aluno alterar no primeiro acesso."""
    alfabeto = string.ascii_letters + string.digits + "!@#$"
    return "".join(secrets.choice(alfabeto) for _ in range(12))


def _normalizar_headers(headers: list[str]) -> dict[str, str]:
    """
    Recebe a lista de cabeçalhos do arquivo e retorna um dict
    { campo_interno: nome_real_no_arquivo } para acesso normalizado.
    """
    mapa = {}
    headers_lower = {h.strip().lower(): h for h in headers if h}
    for campo, sinonimos in MAPA_COLUNAS.items():
        for sin in sinonimos:
            if sin in headers_lower:
                mapa[campo] = headers_lower[sin]
                break
    return mapa


def _validar_linha(
    idx: int,
    row: dict[str, Any],
    mapa: dict[str, str],
) -> dict[str, Any]:
    """
    Valida uma linha do arquivo e retorna um dict com os dados e lista de erros.
    idx começa em 0; linha no relatório = idx + 2 (1 do header + 1 base).
    """
    numero_linha = idx + 2
    erros = []

    # Extrai valores usando o mapa de colunas
    nome  = str(row.get(mapa.get("nome",  ""), "") or "").strip()
    email = str(row.get(mapa.get("email", ""), "") or "").strip().lower()
    nivel = str(row.get(mapa.get("nivel", ""), "") or "").strip().upper()
    serie = str(row.get(mapa.get("serie", ""), "") or "").strip()

    # Validações
    if not nome or len(nome) < 2:
        erros.append("Nome inválido (mínimo 2 caracteres).")

    if not email or "@" not in email or "." not in email.split("@")[-1]:
        erros.append("E-mail inválido.")

    if not nivel:
        erros.append("Nível obrigatório.")
    elif nivel not in NIVEIS_VALIDOS:
        erros.append(
            f"Nível '{nivel}' inválido. "
            f"Use: {', '.join(sorted(NIVEIS_VALIDOS))}."
        )

    return {
        "linha": numero_linha,
        "nome": nome,
        "email": email,
        "nivel": nivel,
        "serie": serie or None,
        "erros": erros,
        "valido": len(erros) == 0,
    }

# Parsers de arquivo

def _ler_csv(conteudo: bytes) -> tuple[list[str], list[dict]]:
    """Retorna (headers, lista_de_dicts)."""
    # Tenta UTF-8 com BOM, depois latin-1
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = conteudo.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(
            status_code=400,
            detail="Não foi possível decodificar o arquivo CSV. Use UTF-8 ou Latin-1.",
        )

    reader = csv.DictReader(io.StringIO(texto))
    headers = reader.fieldnames or []
    linhas = list(reader)
    return list(headers), linhas


def _ler_excel(conteudo: bytes) -> tuple[list[str], list[dict]]:
    """Retorna (headers, lista_de_dicts) lendo a primeira aba."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Arquivo Excel inválido ou corrompido.",
        )

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia.")

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    linhas = []
    for row in rows[1:]:
        # Pula linhas completamente vazias
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        linhas.append(dict(zip(headers, [str(c).strip() if c is not None else "" for c in row])))

    return headers, linhas

# Função principal

async def importar_alunos(
    arquivo: UploadFile,
    db: Session,
) -> dict:
    """
    US33: lê o arquivo, valida cada linha e insere os alunos válidos.

    Retorna relatório com:
    - total_linhas, total_importados, total_erros, total_duplicados
    - lista detalhada de erros e duplicatas
    - lista de alunos importados com e-mail e senha provisória
    """

    # --- Validação do arquivo ---
    if arquivo.content_type not in CONTENT_TYPES_PERMITIDOS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Tipo de arquivo não suportado: '{arquivo.content_type}'. "
                "Envie um arquivo .csv ou .xlsx."
            ),
        )

    conteudo = await arquivo.read()

    if len(conteudo) == 0:
        raise HTTPException(status_code=400, detail="O arquivo está vazio.")

    if len(conteudo) > TAMANHO_MAXIMO:
        raise HTTPException(
            status_code=400,
            detail=f"Arquivo muito grande. Máximo permitido: 5 MB.",
        )

    # --- Leitura do arquivo ---
    nome_lower = arquivo.filename.lower() if arquivo.filename else ""
    if nome_lower.endswith(".xlsx") or nome_lower.endswith(".xls"):
        headers, linhas = _ler_excel(conteudo)
    else:
        headers, linhas = _ler_csv(conteudo)

    if not linhas:
        raise HTTPException(
            status_code=400,
            detail="O arquivo não contém dados (apenas cabeçalho ou vazio).",
        )

    # --- Valida cabeçalhos mínimos ---
    mapa = _normalizar_headers(headers)
    campos_obrigatorios = ["nome", "email", "nivel"]
    faltando = [c for c in campos_obrigatorios if c not in mapa]
    if faltando:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Colunas obrigatórias não encontradas: {', '.join(faltando)}. "
                f"Cabeçalhos detectados: {', '.join(headers)}. "
                "Verifique se o arquivo usa os nomes: nome, email, nivel, serie."
            ),
        )

    # --- Pré-carrega e-mails já existentes no banco (evita N queries) ---
    emails_existentes = {
        u.email.lower()
        for u in db.query(models.Usuario.email).all()
    }

    # --- Processa cada linha ---
    erros: list[dict] = []
    duplicatas: list[dict] = []
    importados: list[dict] = []
    emails_nesta_importacao: set[str] = set()  # evita duplicatas dentro do próprio arquivo

    for idx, row in enumerate(linhas):
        resultado = _validar_linha(idx, row, mapa)

        if not resultado["valido"]:
            erros.append({
                "linha": resultado["linha"],
                "dados": {"nome": resultado["nome"], "email": resultado["email"]},
                "erros": resultado["erros"],
            })
            continue

        email = resultado["email"]

        # Duplicata no banco?
        if email in emails_existentes:
            duplicatas.append({
                "linha": resultado["linha"],
                "email": email,
                "motivo": "E-mail já cadastrado no sistema.",
            })
            continue

        # Duplicata no próprio arquivo?
        if email in emails_nesta_importacao:
            duplicatas.append({
                "linha": resultado["linha"],
                "email": email,
                "motivo": "E-mail duplicado no arquivo (linha anterior já processada).",
            })
            continue

        # Tudo certo — cria o usuário
        senha_provisoria = _gerar_senha_provisoria()
        novo_usuario = models.Usuario(
            nome=resultado["nome"],
            email=email,
            senha_hash=hash_senha(senha_provisoria),
            perfil="ALUNO",
            nivel=resultado["nivel"],
            serie=resultado["serie"],
            status="ATIVO",
        )
        db.add(novo_usuario)
        emails_nesta_importacao.add(email)
        emails_existentes.add(email)  # evita duplicata se o arquivo repete

        importados.append({
            "linha": resultado["linha"],
            "nome": resultado["nome"],
            "email": email,
            "nivel": resultado["nivel"],
            "serie": resultado["serie"],
            # Senha provisória retornada apenas neste momento — admin deve comunicar ao aluno
            "senha_provisoria": senha_provisoria,
        })

    # Commit em lote — mais eficiente que commit por linha
    if importados:
        try:
            db.commit()
        except Exception as exc:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Erro ao salvar alunos no banco: {str(exc)}",
            )

    return {
        "total_linhas": len(linhas),
        "total_importados": len(importados),
        "total_duplicados": len(duplicatas),
        "total_erros": len(erros),
        "importados": importados,
        "duplicados": duplicatas,
        "erros": erros,
    }

# Gerador do arquivo modelo para download

def gerar_modelo_xlsx() -> bytes:
    """
    Gera um arquivo .xlsx de exemplo para o admin baixar e preencher.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alunos"

    # Cabeçalhos
    cabecalhos = ["nome", "email", "nivel", "serie"]
    ws.append(cabecalhos)

    # Estilo do cabeçalho
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0B57C5", end_color="0B57C5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Exemplos de dados
    exemplos = [
        ["João Silva",   "joao.silva@escola.edu.br",   "MEDIO",          "3"],
        ["Maria Santos", "maria.santos@escola.edu.br",  "FUNDAMENTAL_II", "8"],
        ["Pedro Costa",  "pedro.costa@escola.edu.br",   "EJA",            ""],
        ["Ana Lima",     "ana.lima@escola.edu.br",       "FUNDAMENTAL_I",  "4"],
        ["Carlos Souza", "carlos.souza@escola.edu.br",  "ENEM",           ""],
    ]
    for row in exemplos:
        ws.append(row)

    # Largura das colunas
    for col, width in zip(["A", "B", "C", "D"], [30, 35, 18, 10]):
        ws.column_dimensions[col].width = width

    # Aba de instruções
    ws_inst = wb.create_sheet("Instruções")
    instrucoes = [
        ["Campo",  "Obrigatório", "Valores aceitos"],
        ["nome",   "Sim",         "Nome completo (mín. 2 caracteres)"],
        ["email",  "Sim",         "E-mail válido e único"],
        ["nivel",  "Sim",         "FUNDAMENTAL_I | FUNDAMENTAL_II | MEDIO | ENEM | EJA"],
        ["serie",  "Não",         "Ex: 1, 2, 3, 1º ano, etc."],
    ]
    for row in instrucoes:
        ws_inst.append(row)
    for col, width in zip(["A", "B", "C"], [15, 14, 55]):
        ws_inst.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()