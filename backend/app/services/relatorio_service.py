from typing import Optional, List, Dict, Any
from datetime import datetime
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from app import models


def obter_tentativas_filtradas(
    db: Session,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    prova_id: Optional[int] = None,
    nivel: Optional[str] = None,
    serie: Optional[str] = None,
    componente_id: Optional[int] = None,
) -> List[models.Tentativa]:
    """Retorna tentativas com status CONCLUIDA aplicando filtros."""

    query = (
        db.query(models.Tentativa)
        # Carrega aluno e prova junto (joinedload) para evitar N+1 nas agregações
        .options(
            joinedload(models.Tentativa.aluno),
            joinedload(models.Tentativa.prova),
        )
        .join(models.Usuario, models.Tentativa.aluno_id == models.Usuario.id)
        .join(models.Prova, models.Tentativa.prova_id == models.Prova.id)
        .filter(models.Tentativa.status == "CONCLUIDA")
    )

    if data_inicio:
        query = query.filter(models.Tentativa.data_inicio >= data_inicio)
    if data_fim:
        query = query.filter(models.Tentativa.data_fim <= data_fim)
    if prova_id:
        query = query.filter(models.Tentativa.prova_id == prova_id)
    if nivel:
        query = query.filter(models.Usuario.nivel == nivel)
    if serie:
        query = query.filter(models.Usuario.serie == serie)
    if componente_id:
        componente = db.query(models.ComponenteCurricular).filter(
            models.ComponenteCurricular.id == componente_id
        ).first()
        if componente:
            if componente.nivel:
                query = query.filter(models.Prova.nivel == componente.nivel)
            if componente.serie:
                query = query.filter(models.Prova.serie == componente.serie)
    return query.all()


def calcular_estatisticas_gerais(tentativas: List[models.Tentativa]) -> Dict[str, Any]:
    """Calcula média, aprovação e taxa geral."""
    if not tentativas:
        return {
            "total_tentativas": 0,
            "media_geral": 0.0,
            "aprovados": 0,
            "reprovados": 0,
            "taxa_aprovacao_percentual": 0.0,
        }

    total = len(tentativas)
    notas = [float(t.nota) for t in tentativas if t.nota is not None]
    media = sum(notas) / len(notas) if notas else 0.0
    aprovados = sum(1 for t in tentativas if t.resultado == "APROVADO")

    return {
        "total_tentativas": total,
        "media_geral": round(media, 2),
        "aprovados": aprovados,
        "reprovados": total - aprovados,
        "taxa_aprovacao_percentual": round(aprovados / total * 100, 2),
    }


def calcular_por_nivel(tentativas: List[models.Tentativa]) -> List[Dict[str, Any]]:
    """Agrupa e calcula estatísticas por nível de escolaridade."""
    niveis = set(t.aluno.nivel for t in tentativas if t.aluno.nivel)
    resultado = []

    for nivel in sorted(niveis):
        grupo = [t for t in tentativas if t.aluno.nivel == nivel]
        notas = [float(t.nota) for t in grupo if t.nota is not None]
        aprovados = sum(1 for t in grupo if t.resultado == "APROVADO")

        resultado.append({
            "nivel": nivel,
            "total_alunos": len(set(t.aluno_id for t in grupo)),
            "total_tentativas": len(grupo),
            "media_notas": round(sum(notas) / len(notas), 2) if notas else 0.0,
            "taxa_aprovacao_percentual": round(aprovados / len(grupo) * 100, 2),
        })

    return resultado


def calcular_por_serie(tentativas: List[models.Tentativa]) -> List[Dict[str, Any]]:
    """Agrupa e calcula estatísticas por série."""
    series = set(t.aluno.serie for t in tentativas if t.aluno.serie)
    resultado = []

    for serie in sorted(series):
        grupo = [t for t in tentativas if t.aluno.serie == serie]
        notas = [float(t.nota) for t in grupo if t.nota is not None]
        aprovados = sum(1 for t in grupo if t.resultado == "APROVADO")

        resultado.append({
            "serie": serie,
            "total_alunos": len(set(t.aluno_id for t in grupo)),
            "total_tentativas": len(grupo),
            "media_notas": round(sum(notas) / len(notas), 2) if notas else 0.0,
            "taxa_aprovacao_percentual": round(aprovados / len(grupo) * 100, 2),
        })

    return resultado


def calcular_por_componente(
    db: Session,
    tentativas: List[models.Tentativa],
) -> List[Dict[str, Any]]:
    """Agrupa e calcula estatísticas por componente curricular."""
    componentes = db.query(models.ComponenteCurricular).all()
    resultado = []

    for comp in componentes:
        # Filtra tentativas cujas provas são do nível/série do componente
        grupo = [
            t for t in tentativas
            if (not comp.nivel or t.prova.nivel == comp.nivel)
            and (not comp.serie or t.prova.serie == comp.serie)
        ]

        if not grupo:
            continue

        notas = [float(t.nota) for t in grupo if t.nota is not None]
        aprovados = sum(1 for t in grupo if t.resultado == "APROVADO")

        resultado.append({
            "componente_id": comp.id,
            "componente_nome": comp.nome,
            "total_provas": len(set(t.prova_id for t in grupo)),
            "total_tentativas": len(grupo),
            "media_notas": round(sum(notas) / len(notas), 2) if notas else 0.0,
            "taxa_aprovacao_percentual": round(aprovados / len(grupo) * 100, 2),
        })

    return resultado


def calcular_detalhes_provas(
    db: Session,
    tentativas: List[models.Tentativa],
) -> List[Dict[str, Any]]:
    """Calcula estatísticas por prova individual."""
    resultado = []
    provas_ids = set(t.prova_id for t in tentativas)

    for prova_id in provas_ids:
        prova = db.query(models.Prova).filter(models.Prova.id == prova_id).first()
        grupo = [t for t in tentativas if t.prova_id == prova_id]

        if not prova or not grupo:
            continue

        notas = [float(t.nota) for t in grupo if t.nota is not None]
        aprovados = sum(1 for t in grupo if t.resultado == "APROVADO")

        respostas = db.query(models.Resposta).filter(
            models.Resposta.tentativa_id.in_([t.id for t in grupo])
        ).all()
        acertos = sum(1 for r in respostas if r.is_correta)

        resultado.append({
            "prova_id": prova.id,
            "prova_titulo": prova.titulo,
            "tipo": prova.tipo,
            "nivel": prova.nivel,
            "serie": prova.serie,
            "total_tentativas": len(grupo),
            "media_notas": round(sum(notas) / len(notas), 2) if notas else 0.0,
            "acertos_totais": acertos,
            "erros_totais": len(respostas) - acertos,
            "taxa_aprovacao_percentual": round(aprovados / len(grupo) * 100, 2),
            "nota_minima": float(prova.nota_minima or 0),
        })

    return sorted(resultado, key=lambda x: x["media_notas"], reverse=True)


def obter_dados_exportacao(
    db: Session,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    prova_id: Optional[int] = None,
    nivel: Optional[str] = None,
    serie: Optional[str] = None,
    componente_id: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """Retorna dados formatados para exportação."""
    tentativas = obter_tentativas_filtradas(
        db, data_inicio, data_fim, prova_id, nivel, serie, componente_id
    )

    return [
        {
            "aluno": t.aluno.nome,
            "email": t.aluno.email,
            "nivel": t.aluno.nivel or "-",
            "serie": t.aluno.serie or "-",
            "prova": t.prova.titulo,
            "tipo_prova": t.prova.tipo,
            "nota": f"{t.nota:.2f}" if t.nota is not None else "-",
            "resultado": t.resultado or "-",
            "data_inicio": t.data_inicio.strftime("%d/%m/%Y %H:%M") if t.data_inicio else "-",
            "data_conclusao": t.data_fim.strftime("%d/%m/%Y %H:%M") if t.data_fim else "-",
        }
        for t in tentativas
    ]


def exportar_excel(dados: List[Dict[str, Any]]) -> bytes:
    """Gera arquivo Excel formatado."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    headers = [
        "Aluno", "Email", "Nível", "Série", "Prova", "Tipo",
        "Nota", "Resultado", "Data Início", "Data Conclusão"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for dado in dados:
        ws.append([
            dado["aluno"], dado["email"], dado["nivel"], dado["serie"],
            dado["prova"], dado["tipo_prova"], dado["nota"],
            dado["resultado"], dado["data_inicio"], dado["data_conclusao"],
        ])

    for col, width in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"],
        [25,  30,  15,  10,  35,  15,  10,  12,  18,  18]
    ):
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def exportar_csv(dados: List[Dict[str, Any]]) -> bytes:
    """Gera arquivo CSV com encoding UTF-8."""
    import csv
    import io
    from io import BytesIO

    text_stream = io.StringIO()
    fieldnames = [
        "aluno", "email", "nivel", "serie", "prova", "tipo_prova",
        "nota", "resultado", "data_inicio", "data_conclusao"
    ]
    writer = csv.DictWriter(text_stream, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(dados)

    output = BytesIO(text_stream.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return output.getvalue()